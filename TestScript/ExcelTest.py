import openpyxl
filepath = "D:\测试\TestExcel\model.xlsx"
wb = openpyxl.load_workbook(filepath)

print(wb.worksheets)
ws =wb.worksheets[0]
# 获得激活的worksheet
# ws = wb.active

print(ws)

# 数据可以直接赋值给单元格
ws['A1'] = 42
print(ws['A1'])
d = ws.cell(row=4, column=2, value=10)
# 也可以按行添加
ws.append([1, 2, 3])

# Python类型将自动转换
import datetime
ws['A2'] = datetime.datetime.now()

# 保存文件
wb.save(filepath)
sheet1_obj = wb['Sheet3']

print(sheet1_obj['A1'].value)


# ##### 工作表常用操作
# print(wb.active)  # 获取电子表格Worksheet是否有数据
# print(wb.read_only)  # 是否是以只读方式打开
# print(wb.encoding)  # 获取电子表格的编码
# print('', wb.properties)  # 获取电子表格属性如：标题、作者、创建时间等
# print(wb.worksheets)  # 获取工作表名
# print(wb.get_sheet_names())  # 获取工作表的所有名字
# print(wb.sheetnames)  # 获取工作表的所有名字跟wb.get_sheet_names()一样的功能
# print(wb.get_sheet_by_name('Sheet1'))  # 通过工作表的名字，获取Worksheet对象操作电子表格
# print(wb.create_sheet('python创建的工作表'))  # 创建的工作表,记得用save保存，才保存到硬盘上
# print(wb.copy_worksheet(wb['Sheet1']))  # 复制工作表
#
# #### 工作表的常用操作
# sheet1_obj = wb['Sheet1']
# print(sheet1_obj.title)  # 工作表的标题
# print(sheet1_obj.dimensions)  # 获取表格大小，返回格式如：A1:D6
# print(sheet1_obj.max_row)  # 表格最大行数
# print(sheet1_obj.min_row)  # 表格最小行数
# print(sheet1_obj.max_column)  # 表格最大列数
# print(sheet1_obj.min_column)  # 表格最小列数
# print("............")
# print(sheet1_obj.rows)  # 按行获取单元格(Cell对象)
# print(sheet1_obj.columns)  # 按列获取单元格(Cell对象)
# print(sheet1_obj.freeze_panes)  # 冻结窗格
# print(sheet1_obj.values)  # 按行获取表格的内容（数据）
# print(sheet1_obj.iter_rows())#迭代器方式，按行获取所有单元格(Cell对象)
# # print(sheet1_obj.iter_columns())#迭代器方式，按列获取所有单元格(Cell对象)
# # sheet1_obj.append(['1列','2列','3列','4列']) #往工作表最后一行插入多列数据
#
# #### 单元格的常用操作
# # sheet1_obj.merged_cells #合并单元格
# # sheet1_obj.unmerge_cells #取消合并单元格
# print(sheet1_obj['A2'].row)  # 获取行数
# print(sheet1_obj['A2'].column)  # 获取列数
# print(sheet1_obj['B1'].value) #获取单元格的值
# wb.save(filepath) #保存单元格